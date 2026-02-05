#!/usr/bin/env python3
"""
run_demo_from_excel.py

Przykład: czyta dane z Excela i generuje DOCX przez lokalny bot (render-service).
1) POST /render -> rendered.docx
2) POST /replace-image -> final.docx (podmiana placeholdera ALT=REPLACE_ME)

Wymagania:
  pip install openpyxl requests
Uruchom:
  python run_demo_from_excel.py

Założenia:
- render-service działa na http://localhost:3001
- w tym samym katalogu są:
    AAEPW_demo_template.docx
    AAEPW_demo_data.xlsx
    demo_logo_new.png (obraz do podmiany)
Skrypt uruchamia też lokalny serwer HTTP, żeby obraz był dostępny po URL z Excela.
"""

import json
import os
import threading
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler

import requests
from openpyxl import load_workbook

BOT_URL = os.environ.get("BOT_URL", "http://localhost:3001")
PORT_IMG = int(os.environ.get("IMG_PORT", "8000"))

TEMPLATE = "AAEPW_demo_template.docx"
XLSX = "AAEPW_demo_data.xlsx"

OUT_RENDERED = "out_rendered.docx"
OUT_FINAL = "out_final.docx"


def start_static_server(port: int):
    handler = SimpleHTTPRequestHandler
    httpd = ThreadingHTTPServer(("127.0.0.1", port), handler)
    t = threading.Thread(target=httpd.serve_forever, daemon=True)
    t.start()
    return httpd


def read_excel(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)

    main = wb["Main"]
    headers = [c.value for c in main[1]]
    values = [c.value for c in main[2]]
    main_dict = dict(zip(headers, values))

    rows_ws = wb["Rows"]
    rows = []
    rows_headers = [c.value for c in rows_ws[1]]
    for r in rows_ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(rows_headers, r)))

    data = {
        "short_text": str(main_dict.get("short_text", "")),
        "word1": str(main_dict.get("word1", "")),
        "word2": str(main_dict.get("word2", "")),
        "word3": str(main_dict.get("word3", "")),
        "word4": str(main_dict.get("word4", "")),
        "rows": rows,
    }
    obraz_url = str(main_dict.get("obraz_url", "")).strip()
    return data, obraz_url


def render_docx(template_path: str, data: dict, out_path: str):
    with open(template_path, "rb") as f:
        files = {"template": ("template.docx", f, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")}
        resp = requests.post(
            f"{BOT_URL}/render",
            files=files,
            data={"data": json.dumps(data, ensure_ascii=False)},
            timeout=30,
        )
    resp.raise_for_status()
    with open(out_path, "wb") as out:
        out.write(resp.content)


def replace_image(docx_path: str, obraz_url: str, out_path: str):
    with open(docx_path, "rb") as f:
        files = {"docx": ("in.docx", f, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")}
        resp = requests.post(
            f"{BOT_URL}/replace-image",
            files=files,
            data={"data": json.dumps({"obraz_url": obraz_url}, ensure_ascii=False)},
            timeout=30,
        )
    resp.raise_for_status()
    with open(out_path, "wb") as out:
        out.write(resp.content)


def main():
    # uruchom lokalny serwer obrazów (żeby URL z Excela działał)
    httpd = start_static_server(PORT_IMG)
    print(f"[OK] Local image server: http://127.0.0.1:{PORT_IMG}/ (serwuje bieżący katalog)")

    data, obraz_url = read_excel(XLSX)
    print("[OK] Excel loaded")
    print("     obraz_url =", obraz_url)

    print("[..] Rendering DOCX...")
    render_docx(TEMPLATE, data, OUT_RENDERED)
    print(f"[OK] Saved: {OUT_RENDERED}")

    if obraz_url:
        print("[..] Replacing image...")
        replace_image(OUT_RENDERED, obraz_url, OUT_FINAL)
        print(f"[OK] Saved: {OUT_FINAL}")
    else:
        print("[SKIP] obraz_url empty -> no replace-image call")

    # zostawiamy serwer obrazów w tle na czas działania skryptu
    httpd.shutdown()
    print("[DONE]")


if __name__ == "__main__":
    main()
